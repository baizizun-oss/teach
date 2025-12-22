# sangao/AnswerService.py

import myportal.common as common
import logging
import os
from openpyxl import load_workbook
import config

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)


def normalize_code(code):
    """
    标准化代码内容，去除空白字符差异
    """
    if not code:
        return ""
    return " ".join(code.split())


class SingleChoiceScorer:
    @staticmethod
    def calculateScore(student_answer, correct_answer):
        score = 0
        is_correct = 0
        if str(student_answer).strip() == str(correct_answer).strip():
            score = 2
            is_correct = 1
        return score, is_correct


class MultipleChoiceScorer:
    @staticmethod
    def calculateScore(student_answer, correct_answer):
        score = 0
        is_correct = 0

        def sort_choices(choice_string):
            return ','.join(sorted([x.strip().upper() for x in choice_string.split(',') if x.strip()])) if choice_string else ""

        if sort_choices(student_answer) == sort_choices(correct_answer):
            score = 4
            is_correct = 1
        return score, is_correct


class FillBlankScorer:
    @staticmethod
    def calculateScore(student_answer, correct_answer):
        score = 0
        is_correct = 0
        acceptable_options = [option.strip().lower() for option in correct_answer.split('|')]
        if student_answer.strip().lower() in acceptable_options:
            score = 1
            is_correct = 1
        return score, is_correct


class TrueFalseScorer:
    @staticmethod
    def calculateScore(student_answer, correct_answer):
        score = 0
        is_correct = 0
        if str(student_answer).strip() == str(correct_answer).strip():
            score = 2
            is_correct = 1
        return score, is_correct


def get_submission_id():
    import uuid
    return str(uuid.uuid4())


def get_upload_filepath(file, user_id, question_id):
    import time
    file_obj = file
    logger.info(f"file: {file_obj}")
    filename = file_obj["filename"]
    body = file_obj["body"]

    upload_dir = config.get_path("sangao", "Answer", "files")
    os.makedirs(upload_dir, exist_ok=True)

    timestamp = str(int(time.time()))
    safe_filename = f"{user_id}_{question_id}_{timestamp}_{filename}"
    file_path = os.path.join(upload_dir, safe_filename)

    with open(file_path, "wb") as f:
        f.write(body)

    logger.info(f"Saved file to: {file_path}")
    return safe_filename



import myportal.common as common
import logging
import os
import re
import zipfile
from openpyxl import load_workbook
from openpyxl.styles import Color
from openpyxl.cell.cell import Cell   # <-- 添加这一行
import config
import json

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)


def normalize_value(v):
    """标准化值用于比对"""
    if v is None:
        return ""
    if isinstance(v, float):
        return round(v, 10)  # 消除浮点误差
    return str(v).strip()


def normalize_formula(f):
    """标准化公式：去空格、去$、转大写"""
    if not isinstance(f, str) or not f.startswith('='):
        return ""
    return f.replace(" ", "").replace("$", "").upper()


def get_cell_color(cell):
    """获取单元格填充色的 RGB 值（字符串），若无则返回 None"""
    fill = cell.fill
    if fill and fill.start_color and fill.start_color.type == 'rgb':
        return fill.start_color.rgb
    return None


class OperationScorer:
    def __init__(self, student_file_path, correct_answer_file_path, scoring_rules):
        self.student_file_path = student_file_path
        self.correct_answer_file_path = correct_answer_file_path
        self.scoring_rules = scoring_rules  # 必须是 dict，如 {"cell_values": {"desc": "...", "max_score": 5}, ...}
        logger.info(f"scoring_rules:{self.scoring_rules}")
        self.compare_range = scoring_rules["compare_range"]
        self.ws_student_val = load_workbook(os.path.join(config.get_path("sangao","Answer","files"),student_file_path), data_only=True).active
        self.ws_student_form = load_workbook(os.path.join(config.get_path("sangao","Answer","files"),student_file_path), data_only=False).active
        self.ws_correct_val = load_workbook(os.path.join(config.get_path("sangao","Question","files","operation"),correct_answer_file_path), data_only=True).active
        self.ws_correct_form = load_workbook(os.path.join(config.get_path("sangao","Question","files","operation"),correct_answer_file_path), data_only=False).active

    def score(self):
        # 在 OperationScorer.score() 开头
        scoring_details = {}
        total_max = 0

        for key, rule in self.scoring_rules.items():
            if isinstance(rule, dict) and "desc" in rule and "max_score" in rule:
                scoring_details[key] = {
                    "desc": rule["desc"],
                    "score": 0,
                    "max_score": rule["max_score"]
                }
                total_max += rule["max_score"]

        scoring_details["total_score"] = 0
        scoring_details["max_total_score"] = total_max        




        try:
            if scoring_details.get("merged_cells", {}).get("max_score", 0) > 0:
                merged_ok = set(self.ws_student_val.merged_cells.ranges) == set(self.ws_correct_val.merged_cells.ranges)
                scoring_details["merged_cells"]["score"] = scoring_details["merged_cells"]["max_score"] if merged_ok else 0

            if scoring_details.get("cell_values", {}).get("max_score", 0) > 0:
                match, total = self._compare_cell_values()
                ratio = match / total if total > 0 else 0
                scoring_details["cell_values"]["score"] = round(ratio * scoring_details["cell_values"]["max_score"], 1)

            if scoring_details.get("formulas", {}).get("max_score", 0) > 0:
                match, total = self._compare_formulas()
                ratio = match / total if total > 0 else 0
                scoring_details["formulas"]["score"] = round(ratio * scoring_details["formulas"]["max_score"], 1)

            if scoring_details.get("conditional_formatting", {}).get("max_score", 0) > 0:
                color_ok = self._compare_conditional_formatting_by_color()
                scoring_details["conditional_formatting"]["score"] = scoring_details["conditional_formatting"]["max_score"] if color_ok else 0

            if scoring_details.get("chart", {}).get("max_score", 0) > 0:
                chart_ok = self._compare_charts()
                scoring_details["chart"]["score"] = scoring_details["chart"]["max_score"] if chart_ok else 0

            total_score = sum(detail["score"] for key, detail in scoring_details.items() 
                             if key not in ("total_score", "max_total_score"))
            scoring_details["total_score"] = round(total_score, 1)

            return total_score, scoring_details

        except Exception as e:
            logger.error(f"通用评分出错: {e}", exc_info=True)
            return 0, scoring_details

    def _compare_cell_values(self):
        match = 0
        total = 0

        # 如果指定了比较范围，就只比对那个区域
        if self.compare_range:
            cells = self.ws_correct_val[self.compare_range]
            if isinstance(cells, Cell):  # 单个单元格
                cells = [[cells]]
            for row in cells:
                for cell in row:
                    if cell.value is None:
                        continue
                    total += 1
                    coord = cell.coordinate
                    student_cell = self.ws_student_val[coord]
                    if normalize_value(student_cell.value) == normalize_value(cell.value):
                        match += 1
        else:
            # 全表比对（原逻辑）
            for row in self.ws_correct_val.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    total += 1
                    coord = cell.coordinate
                    student_cell = self.ws_student_val[coord]
                    if normalize_value(student_cell.value) == normalize_value(cell.value):
                        match += 1

        return match, total

    def _compare_formulas(self):
        match = 0
        total = 0
        for row in self.ws_correct_form.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    total += 1
                    coord = cell.coordinate
                    student_cell = self.ws_student_form[coord]
                    if normalize_formula(student_cell.value) == normalize_formula(cell.value):
                        match += 1
        return match, total

    def _compare_conditional_formatting_by_color(self):
        DEFAULT_COLOR = "00000000"
        colored_cells = {}
        for row in self.ws_correct_val.iter_rows():
            for cell in row:
                color = get_cell_color(cell)
                if color and color not in (DEFAULT_COLOR, "FFFFFFFF"):
                    colored_cells[cell.coordinate] = color

        if not colored_cells:
            return True

        for coord, expected_color in colored_cells.items():
            student_cell = self.ws_student_val[coord]
            actual_color = get_cell_color(student_cell)
            if actual_color != expected_color:
                return False
        return True

    def _compare_charts(self):
        try:
            correct_features = self._extract_chart_features(self.correct_answer_file_path)
            student_features = self._extract_chart_features(self.student_file_path)
            return self._charts_match(correct_features, student_features)
        except Exception as e:
            logger.warning(f"图表比对失败: {e}")
            return False



    def _extract_chart_features(self, file_path):
        features = []
        with zipfile.ZipFile(file_path, 'r') as zf:
            chart_files = [f for f in zf.namelist() if f.startswith('xl/charts/chart') and f.endswith('.xml')]
            if not chart_files:
                return {}

            content = zf.read(chart_files[0]).decode('utf-8', errors='ignore')
            chart_type = "line" if "<c:lineChart>" in content else \
                         "bar" if "<c:barChart>" in content or "<c:colChart>" in content else "unknown"

            refs = re.findall(r'<c:f>([^<]+)</c:f>', content)
            x_ref = None
            y_refs = []
            for ref in refs:
                if '$A$' in ref or 'A3' in ref or 'A4' in ref:
                    x_ref = ref
                elif any(col in ref for col in ['$B$', '$C$', '$D$', '$E$', '$F$', '$G$', '$H$', '$I$']):
                    y_refs.append(ref)

            title_match = re.search(r'<c:title>.*?<c:val val="([^"]+)"', content)
            title = title_match.group(1) if title_match else ""

            return {
                "type": chart_type,
                "x_ref": x_ref,
                "y_refs": y_refs,
                "title": title
            }

    def _charts_match(self, correct, student):
        if not correct or not student:
            return correct == student
        if correct["type"] != student["type"]:
            return False

        # 饼图特殊处理
        if correct["type"] == "pie":
            # 饼图至少要有 Y 数据（值），标题尽量匹配
            has_y = len(student.get("y_refs", [])) >= 1
            title_ok = True
            if correct["title"].strip():
                title_ok = correct["title"].lower() in student.get("title", "").lower()
            return has_y and title_ok

        # 原有逻辑用于 line/bar
        if correct["x_ref"] and not (student["x_ref"] and 'A' in student["x_ref"]):
            return False
        if len(correct["y_refs"]) != len(student["y_refs"]):
            return False
        if correct["title"] and correct["title"].strip():
            if correct["title"].lower() not in student["title"].lower():
                return False
        return True


# ========================
# 统一的操作题评分入口
# ========================
def calculateScore(submitted_file_path, question_id, correct_answer_file_path):
    """
    使用 OperationScorer 对操作题进行评分。
    评分规则从数据库 OperationQuestionModel.score_rules 字段读取（必须是 JSON 字典）。
    :param submitted_file_path: 学生提交的 Excel 文件路径
    :param question_id: 题目 ID
    :param correct_answer_file_path: 标准答案 Excel 文件路径
    :return: (total_score, scoring_details_dict)
    """
    from common.OperationQuestionModel import OperationQuestionModel

    try:
        question_model = OperationQuestionModel(question_id)
        score_rules_str = question_model.score_rules

        if not score_rules_str or not score_rules_str.strip():
            logger.warning(f"题目 {question_id} 未配置评分规则，返回0分")
            empty_details = {
                "total_score": 0,
                "max_total_score": 0,
                "error": {"desc": "未配置评分规则", "score": 0, "max_score": 0}
            }
            return 0, empty_details

        # 解析为字典（不再是列表！）
        scoring_rules = json.loads(score_rules_str)
        logger.info(f"scoring_rules:{scoring_rules}")
        # 安全检查：必须是 dict
        if not isinstance(scoring_rules, dict):
            raise ValueError("score_rules 必须是 JSON 对象（字典），而非数组或字符串")

        # 初始化评分器
        scorer = OperationScorer(
            student_file_path=submitted_file_path,
            correct_answer_file_path=correct_answer_file_path,
            scoring_rules=scoring_rules
        )

        return scorer.score()

    except json.JSONDecodeError as e:
        logger.error(f"题目 {question_id} 的 score_rules 不是合法 JSON: {e}")
        return 0, {"error": {"desc": "评分规则 JSON 格式错误", "score": 0, "max_score": 0}}
    except ValueError as e:
        logger.error(f"题目 {question_id} 的 score_rules 结构错误: {e}")
        return 0, {"error": {"desc": str(e), "score": 0, "max_score": 0}}
    except Exception as e:
        logger.exception(f"评分过程异常: {e}")
        return 0, {"error": {"desc": f"评分异常: {str(e)}", "score": 0, "max_score": 0}}


# 其他辅助函数（如 normalize_code, get_submission_id 等可保留）
def normalize_code(code):
    if not code:
        return ""
    return " ".join(code.split())


def get_submission_id():
    import uuid
    return str(uuid.uuid4())


def get_upload_filepath(file, user_id, question_id):
    import time
    file_obj = file
    logger.info(f"file: {file_obj}")
    filename = file_obj["filename"]
    body = file_obj["body"]

    upload_dir = config.get_path("sangao", "Answer", "files")
    os.makedirs(upload_dir, exist_ok=True)

    timestamp = str(int(time.time()))
    safe_filename = f"{user_id}_{question_id}_{timestamp}_{filename}"
    file_path = os.path.join(upload_dir, safe_filename)

    with open(file_path, "wb") as f:
        f.write(body)

    logger.info(f"Saved file to: {file_path}")
    return safe_filename