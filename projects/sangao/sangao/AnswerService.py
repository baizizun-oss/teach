# sangao/AnswerService.py

import myportal.common as common
import logging
import os
import re
import zipfile
from openpyxl import load_workbook
from openpyxl.styles import Color
from openpyxl.cell.cell import Cell  # <-- å·²æ­£ç¡®å¯¼å…¥
import config
import json

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)


def normalize_value(v):
    """æ ‡å‡†åŒ–å€¼ç”¨äºæ¯”å¯¹"""
    if v is None:
        return ""
    if isinstance(v, float):
        return round(v, 10)
    return str(v).strip()


def normalize_formula(f):
    """æ ‡å‡†åŒ–å…¬å¼ï¼šå»ç©ºæ ¼ã€å»$ã€è½¬å¤§å†™"""
    if not isinstance(f, str) or not f.startswith('='):
        return ""
    return f.replace(" ", "").replace("$", "").upper()


def get_cell_color(cell):
    """è·å–å•å…ƒæ ¼å¡«å……è‰²çš„ RGB å€¼ï¼ˆå­—ç¬¦ä¸²ï¼‰ï¼Œè‹¥æ— åˆ™è¿”å› None"""
    fill = cell.fill
    if fill and fill.start_color and fill.start_color.type == 'rgb':
        return fill.start_color.rgb
    return None


class OperationScorer:
    def __init__(self, student_file_path, correct_answer_file_path, scoring_rules):
        self.student_file_path = student_file_path
        self.correct_answer_file_path = correct_answer_file_path
        self.scoring_rules = scoring_rules
        logger.info(f"scoring_rules:{self.scoring_rules}")
        self.compare_range = scoring_rules.get("compare_range", "")
        self.ws_student_val = load_workbook(
            os.path.join(config.get_path("sangao", "Answer", "files"), student_file_path),
            data_only=True
        ).active
        self.ws_student_form = load_workbook(
            os.path.join(config.get_path("sangao", "Answer", "files"), student_file_path),
            data_only=False
        ).active
        self.ws_correct_val = load_workbook(
            os.path.join(config.get_path("sangao", "Question", "files", "operation"), correct_answer_file_path),
            data_only=True
        ).active
        self.ws_correct_form = load_workbook(
            os.path.join(config.get_path("sangao", "Question", "files", "operation"), correct_answer_file_path),
            data_only=False
        ).active

    def score(self):
        scoring_details = {}
        total_max = 0

        for key, rule in self.scoring_rules.items():
            if isinstance(rule, dict) and "desc" in rule and "max_score" in rule:
                scoring_details[key] = {
                    "desc": rule["desc"],
                    "score": 0,
                    "max_score": rule["max_score"]
                }
                total_max += rule["max_score"]

        scoring_details["total_score"] = 0
        scoring_details["max_total_score"] = total_max

        try:
            if scoring_details.get("merged_cells", {}).get("max_score", 0) > 0:
                merged_ok = set(self.ws_student_val.merged_cells.ranges) == set(self.ws_correct_val.merged_cells.ranges)
                scoring_details["merged_cells"]["score"] = scoring_details["merged_cells"]["max_score"] if merged_ok else 0

            if scoring_details.get("cell_values", {}).get("max_score", 0) > 0:
                match, total = self._compare_cell_values()
                ratio = match / total if total > 0 else 0
                scoring_details["cell_values"]["score"] = round(ratio * scoring_details["cell_values"]["max_score"], 1)

            if scoring_details.get("formulas", {}).get("max_score", 0) > 0:
                match, total = self._compare_formulas()
                ratio = match / total if total > 0 else 0
                scoring_details["formulas"]["score"] = round(ratio * scoring_details["formulas"]["max_score"], 1)

            if scoring_details.get("conditional_formatting", {}).get("max_score", 0) > 0:
                color_ok = self._compare_conditional_formatting_by_color()
                scoring_details["conditional_formatting"]["score"] = scoring_details["conditional_formatting"]["max_score"] if color_ok else 0

            if scoring_details.get("chart", {}).get("max_score", 0) > 0:
                chart_ok = self._compare_charts()
                scoring_details["chart"]["score"] = scoring_details["chart"]["max_score"] if chart_ok else 0

            total_score = sum(
                detail["score"] for key, detail in scoring_details.items()
                if key not in ("total_score", "max_total_score")
            )
            scoring_details["total_score"] = round(total_score, 1)

            return total_score, scoring_details

        except Exception as e:
            logger.error(f"é€šç”¨è¯„åˆ†å‡ºé”™: {e}", exc_info=True)
            return 0, scoring_details

    def _compare_cell_values(self):
        match = 0
        total = 0

        if self.compare_range:
            cells = self.ws_correct_val[self.compare_range]
            if isinstance(cells, Cell):  # å•ä¸ªå•å…ƒæ ¼
                cells = [[cells]]
            for row in cells:
                for cell in row:
                    if cell.value is None:
                        continue
                    total += 1
                    coord = cell.coordinate
                    student_cell = self.ws_student_val[coord]
                    if normalize_value(student_cell.value) == normalize_value(cell.value):
                        match += 1
        else:
            for row in self.ws_correct_val.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    total += 1
                    coord = cell.coordinate
                    student_cell = self.ws_student_val[coord]
                    if normalize_value(student_cell.value) == normalize_value(cell.value):
                        match += 1

        return match, total

    def _compare_formulas(self):
        match = 0
        total = 0
        for row in self.ws_correct_form.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    total += 1
                    coord = cell.coordinate
                    student_cell = self.ws_student_form[coord]
                    if normalize_formula(student_cell.value) == normalize_formula(cell.value):
                        match += 1
        return match, total

    def _compare_conditional_formatting_by_color(self):
        DEFAULT_COLOR = "00000000"
        colored_cells = {}
        for row in self.ws_correct_val.iter_rows():
            for cell in row:
                color = get_cell_color(cell)
                if color and color not in (DEFAULT_COLOR, "FFFFFFFF"):
                    colored_cells[cell.coordinate] = color

        if not colored_cells:
            return True

        for coord, expected_color in colored_cells.items():
            student_cell = self.ws_student_val[coord]
            actual_color = get_cell_color(student_cell)
            if actual_color != expected_color:
                return False
        return True

    def _compare_charts(self):
        try:
            def get_first_chart_xml(file_path):
                with zipfile.ZipFile(file_path, 'r') as zf:
                    chart_files = sorted([f for f in zf.namelist() if f.startswith('xl/charts/chart') and f.endswith('.xml')])
                    if not chart_files:
                        return None
                    content = zf.read(chart_files[0]).decode('utf-8', errors='ignore')
                    # æ ‡å‡†åŒ–ï¼šç§»é™¤æ‰€æœ‰ç©ºç™½å­—ç¬¦ï¼ˆåŒ…æ‹¬æ¢è¡Œã€ç©ºæ ¼ã€åˆ¶è¡¨ç¬¦ï¼‰
                    normalized = re.sub(r'\s+', '', content)
                    return normalized

            correct_xml = get_first_chart_xml(
                os.path.join(config.get_path("sangao", "Question", "files", "operation"), self.correct_answer_file_path)
            )
            student_xml = get_first_chart_xml(
                os.path.join(config.get_path("sangao", "Answer", "files"), self.student_file_path)
            )

            # ä¸¤ç§æƒ…å†µéƒ½æ— å›¾è¡¨ â†’ è§†ä¸ºä¸€è‡´ï¼ˆä½†æœ¬é¢˜åº”æœ‰å›¾è¡¨ï¼‰
            if correct_xml is None and student_xml is None:
                return True
            # ä¸€ä¸ªæœ‰ï¼Œä¸€ä¸ªæ²¡æœ‰ â†’ ä¸ä¸€è‡´
            if correct_xml is None or student_xml is None:
                return False
            # éƒ½æœ‰ â†’ æ¯”å¯¹æ ‡å‡†åŒ–åçš„å†…å®¹
            return correct_xml == student_xml

        except Exception as e:
            logger.warning(f"å›¾è¡¨XMLæ¯”å¯¹å¤±è´¥: {e}")
            return False

    def _extract_chart_features(self, file_path):
        with zipfile.ZipFile(file_path, 'r') as zf:
            chart_files = [f for f in zf.namelist() if f.startswith('xl/charts/chart') and f.endswith('.xml')]
            if not chart_files:
                return {}

            content = zf.read(chart_files[0]).decode('utf-8', errors='ignore')

            # ğŸ”¥ å…³é”®ä¿®æ”¹ï¼šå¢åŠ å¯¹ pieChart çš„è¯†åˆ«
            if "<c:pieChart>" in content:
                chart_type = "pie"
            elif "<c:lineChart>" in content:
                chart_type = "line"
            elif "<c:barChart>" in content or "<c:colChart>" in content:
                chart_type = "bar"
            else:
                chart_type = "unknown"

            refs = re.findall(r'<c:f>([^<]+)</c:f>', content)
            x_ref = None
            y_refs = []
            for ref in refs:
                if '$A$' in ref or 'A3' in ref or 'A4' in ref:
                    x_ref = ref
                elif any(col in ref for col in ['$B$', '$C$', '$D$', '$E$', '$F$', '$G$', '$H$', '$I$']):
                    y_refs.append(ref)

            title_match = re.search(r'<c:title>.*?<c:val val="([^"]+)"', content)
            title = title_match.group(1) if title_match else ""

            return {
                "type": chart_type,
                "x_ref": x_ref,
                "y_refs": y_refs,
                "title": title
            }

    def _charts_match(self, correct, student):
        if not correct or not student:
            return correct == student
        if correct["type"] != student["type"]:
            return False

        # é¥¼å›¾ç‰¹æ®Šå¤„ç†
        if correct["type"] == "pie":
            has_y = len(student.get("y_refs", [])) >= 1
            title_ok = True
            if correct["title"].strip():
                title_ok = correct["title"].lower() in student.get("title", "").lower()
            return has_y and title_ok

        # åŸæœ‰é€»è¾‘ç”¨äº line/bar
        if correct["x_ref"] and not (student["x_ref"] and 'A' in student["x_ref"]):
            return False
        if len(correct["y_refs"]) != len(student["y_refs"]):
            return False
        if correct["title"] and correct["title"].strip():
            if correct["title"].lower() not in student["title"].lower():
                return False
        return True


# ========================
# ç»Ÿä¸€çš„æ“ä½œé¢˜è¯„åˆ†å…¥å£
# ========================
def calculateScore(submitted_file_path, question_id, correct_answer_file_path):
    from common.OperationQuestionModel import OperationQuestionModel

    try:
        question_model = OperationQuestionModel(question_id)
        score_rules_str = question_model.score_rules

        if not score_rules_str or not score_rules_str.strip():
            logger.warning(f"é¢˜ç›® {question_id} æœªé…ç½®è¯„åˆ†è§„åˆ™ï¼Œè¿”å›0åˆ†")
            empty_details = {
                "total_score": 0,
                "max_total_score": 0,
                "error": {"desc": "æœªé…ç½®è¯„åˆ†è§„åˆ™", "score": 0, "max_score": 0}
            }
            return 0, empty_details

        scoring_rules = json.loads(score_rules_str)
        if not isinstance(scoring_rules, dict):
            raise ValueError("score_rules å¿…é¡»æ˜¯ JSON å¯¹è±¡ï¼ˆå­—å…¸ï¼‰")

        scorer = OperationScorer(
            student_file_path=submitted_file_path,
            correct_answer_file_path=correct_answer_file_path,
            scoring_rules=scoring_rules
        )

        return scorer.score()

    except json.JSONDecodeError as e:
        logger.error(f"é¢˜ç›® {question_id} çš„ score_rules ä¸æ˜¯åˆæ³• JSON: {e}")
        return 0, {"error": {"desc": "è¯„åˆ†è§„åˆ™ JSON æ ¼å¼é”™è¯¯", "score": 0, "max_score": 0}}
    except ValueError as e:
        logger.error(f"é¢˜ç›® {question_id} çš„ score_rules ç»“æ„é”™è¯¯: {e}")
        return 0, {"error": {"desc": str(e), "score": 0, "max_score": 0}}
    except Exception as e:
        logger.exception(f"è¯„åˆ†è¿‡ç¨‹å¼‚å¸¸: {e}")
        return 0, {"error": {"desc": f"è¯„åˆ†å¼‚å¸¸: {str(e)}", "score": 0, "max_score": 0}}


# å…¶ä»–è¾…åŠ©å‡½æ•°ï¼ˆä¿æŒä¸å˜ï¼‰
def normalize_code(code):
    if not code:
        return ""
    return " ".join(code.split())


def get_submission_id():
    import uuid
    return str(uuid.uuid4())


def get_upload_filepath(file, user_id, question_id):
    import time
    filename = file["filename"]
    body = file["body"]

    upload_dir = config.get_path("sangao", "Answer", "files")
    os.makedirs(upload_dir, exist_ok=True)

    timestamp = str(int(time.time()))
    safe_filename = f"{user_id}_{question_id}_{timestamp}_{filename}"
    file_path = os.path.join(upload_dir, safe_filename)

    with open(file_path, "wb") as f:
        f.write(body)

    logger.info(f"Saved file to: {file_path}")
    return safe_filename



class SingleChoiceScorer:
    @staticmethod
    def calculateScore(student_answer, correct_answer):
        score = 0
        is_correct = 0
        if str(student_answer).strip() == str(correct_answer).strip():
            score = 2
            is_correct = 1
        return score, is_correct


class MultipleChoiceScorer:
    @staticmethod
    def calculateScore(student_answer, correct_answer):
        score = 0
        is_correct = 0

        def sort_choices(choice_string):
            return ','.join(sorted([x.strip().upper() for x in choice_string.split(',') if x.strip()])) if choice_string else ""

        if sort_choices(student_answer) == sort_choices(correct_answer):
            score = 4
            is_correct = 1
        return score, is_correct


class FillBlankScorer:
    @staticmethod
    def calculateScore(student_answer, correct_answer):
        score = 0
        is_correct = 0
        acceptable_options = [option.strip().lower() for option in correct_answer.split('|')]
        if student_answer.strip().lower() in acceptable_options:
            score = 1
            is_correct = 1
        return score, is_correct


class TrueFalseScorer:
    @staticmethod
    def calculateScore(student_answer, correct_answer):
        score = 0
        is_correct = 0
        if str(student_answer).strip() == str(correct_answer).strip():
            score = 2
            is_correct = 1
        return score, is_correct